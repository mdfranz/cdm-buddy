from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from cdm_wizard.model import (
    FUNCTION_COLORS,
    GOVERN_FUNCTIONS,
    LEFT_OF_BOOM_FUNCTIONS,
    RIGHT_OF_BOOM_FUNCTIONS,
    asset_classes,
    functions,
)


def export_to_excel(data, output_path="cdm_output.xlsx"):
    """
    Exports CDM data to a styled Excel file.
    """
    function_list = functions()
    assets = [asset for asset in asset_classes() if asset in data]

    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cyber Defense Matrix'
    
    # Styles
    header_font = Font(bold=True, size=12)
    boom_font = Font(bold=True, size=14, color="FFFFFF")
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # 1. Add Boom Headers
    govern_col = 2
    left_boom_start = govern_col + len(GOVERN_FUNCTIONS)
    left_boom_end = left_boom_start + len(LEFT_OF_BOOM_FUNCTIONS) - 1
    right_boom_start = left_boom_end + 1
    right_boom_end = right_boom_start + len(RIGHT_OF_BOOM_FUNCTIONS) - 1

    govern_cell = ws.cell(row=1, column=govern_col, value="GOVERN (Cross-Cutting)")
    govern_cell.fill = PatternFill(start_color="6AA84F", end_color="6AA84F", fill_type="solid")
    govern_cell.font = boom_font
    govern_cell.alignment = center_alignment
    govern_cell.border = thin_border

    ws.merge_cells(start_row=1, start_column=left_boom_start, end_row=1, end_column=left_boom_end)
    ws.merge_cells(start_row=1, start_column=right_boom_start, end_row=1, end_column=right_boom_end)

    left_boom_cell = ws.cell(row=1, column=left_boom_start, value="LEFT OF BOOM (Pre-Event)")
    left_boom_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    left_boom_cell.font = boom_font
    left_boom_cell.alignment = center_alignment

    right_boom_cell = ws.cell(row=1, column=right_boom_start, value="RIGHT OF BOOM (Post-Event)")
    right_boom_cell.fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
    right_boom_cell.font = boom_font
    right_boom_cell.alignment = center_alignment

    # 2. Add Function Headers
    for col_idx, func in enumerate(function_list, start=2):
        cell = ws.cell(row=2, column=col_idx, value=func)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = PatternFill(start_color=FUNCTION_COLORS[func], end_color=FUNCTION_COLORS[func], fill_type="solid")
        cell.border = thin_border

    # 3. Add Asset Headers and Data
    for row_idx, asset in enumerate(assets, start=3):
        # Asset Name
        asset_cell = ws.cell(row=row_idx, column=1, value=asset)
        asset_cell.font = header_font
        asset_cell.alignment = center_alignment
        asset_cell.border = thin_border
        
        for col_idx, func in enumerate(function_list, start=2):
            cell_data = data[asset].get(func, {"Tech": "", "People": "", "Process": ""})
            value = f"TECH: {cell_data['Tech']}\nPEOPLE: {cell_data['People']}\nPROCESS: {cell_data['Process']}"
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = left_alignment
            cell.border = thin_border
            cell.fill = PatternFill(start_color=FUNCTION_COLORS[func], end_color=FUNCTION_COLORS[func], fill_type="solid")

    # 4. Add Boom Line (Thick border between Protect and Detect)
    boom_boundary_col = 2 + function_list.index("Protect")
    for r in range(1, len(assets) + 3):
        existing_border = ws.cell(row=r, column=boom_boundary_col).border
        ws.cell(row=r, column=boom_boundary_col).border = Border(
            left=existing_border.left,
            right=Side(style='thick'),
            top=existing_border.top,
            bottom=existing_border.bottom
        )

    # 5. Add Dependency Legend at bottom
    legend_row = len(assets) + 4
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=len(function_list) + 1)
    legend_cell = ws.cell(
        row=legend_row,
        column=1,
        value="Dependency emphasis: Govern is process-led | Identify/Protect lean Technology | Detect/Respond/Recover lean People",
    )
    legend_cell.font = Font(italic=True)
    legend_cell.alignment = center_alignment

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 15
    for col_idx in range(2, len(function_list) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 28
        
    wb.save(output_path)
    return output_path
