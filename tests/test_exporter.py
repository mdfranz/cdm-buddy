import os
import sys
import tempfile
import unittest

from openpyxl import load_workbook


sys.path.append(os.path.join(os.path.dirname(__file__), "../src"))

from cdm_wizard.exporter import export_to_excel
from cdm_wizard.model import empty_matrix, functions


class ExporterTests(unittest.TestCase):
    def build_test_data(self):
        data = empty_matrix()
        data["Devices"]["Govern"] = {
            "Tech": "ServiceNow GRC",
            "People": "Security Governance",
            "Process": "Endpoint security standard",
        }
        data["Devices"]["Identify"] = {
            "Tech": "Lansweeper",
            "People": "IT Ops",
            "Process": "Asset inventory review",
        }
        data["Devices"]["Protect"] = {
            "Tech": "CrowdStrike Falcon",
            "People": "SysAdmin",
            "Process": "Patch Management SOP",
        }
        return data

    def test_export_writes_govern_headers_and_content(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test_cdm.xlsx")
            export_to_excel(self.build_test_data(), output_path)

            wb = load_workbook(output_path)
            ws = wb["Cyber Defense Matrix"]

            self.assertEqual(ws["B1"].value, "GOVERN (Cross-Cutting)")
            self.assertEqual(ws["C1"].value, "LEFT OF BOOM (Pre-Event)")
            self.assertEqual(ws["E1"].value, "RIGHT OF BOOM (Post-Event)")
            self.assertEqual(
                [ws.cell(row=2, column=col).value for col in range(2, 2 + len(functions()))],
                functions(),
            )
            self.assertIn("C1:D1", {str(cell_range) for cell_range in ws.merged_cells.ranges})
            self.assertIn("E1:G1", {str(cell_range) for cell_range in ws.merged_cells.ranges})
            self.assertEqual(ws["D2"].border.right.style, "thick")
            self.assertIn("TECH: ServiceNow GRC", ws["B3"].value)
            self.assertIn("PROCESS: Endpoint security standard", ws["B3"].value)

    def test_export_writes_dependency_legend_for_govern(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test_cdm.xlsx")
            export_to_excel(self.build_test_data(), output_path)

            wb = load_workbook(output_path)
            ws = wb["Cyber Defense Matrix"]
            legend_row = 9

            self.assertIn("Govern is process-led", ws.cell(row=legend_row, column=1).value)


if __name__ == "__main__":
    unittest.main()
